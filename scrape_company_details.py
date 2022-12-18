import logging
import re
import time

import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl import load_workbook
workbook_email = load_workbook("Company_email.xlsx")
only_email = Workbook()
only_email_ws = only_email.active
ws_email = workbook_email.active
ws_email.append(["Company name", "Email"])

contact_workbook = load_workbook("Contact_details.xlsx")
ws_contact = contact_workbook.active
ws_contact.append(["Company name", "Address", "Phone"])

base_url = "https://hwk-kassel.odav.de"
top_limit = 12100
offset= 11800
logging.basicConfig(filename="error.log", level=logging.INFO)


while offset <= top_limit:
    try:
        page = requests.get(
            "https://hwk-kassel.odav.de/betriebe/suche-43,46,bdbsearch.html?limit=12122&search-searchterm=&search-job=&search-local=&search-filter-zipcode=34117&search-filter-latitude=51.315833&search-filter-latitude=51.315833&search-filter-latitude=51.315833&search-filter-latitude=51.315833&search-filter-latitude=51.315833&search-filter-longitude=9.4925&search-filter-longitude=9.4925&search-filter-longitude=9.4925&search-filter-longitude=9.4925&search-filter-longitude=9.4925&search-filter-radius=250&search-filter-jobnr=&offset={}".format(offset))
        soup = BeautifulSoup(page.content, 'html.parser')
    except Exception as e:
        logging.info(e)
        time.sleep(5)
        offset += 100
        continue
    print(f"ofseet count {offset}")
    print("###################")
    all_item_divs = soup.find_all('div', attrs={"class": "searchhit-header"})
    for divs in all_item_divs:
        item_details_url = base_url + divs.find("a").attrs["href"]
        details_page = requests.get(item_details_url)
        details_soup = BeautifulSoup(details_page.content, "html.parser")
        company_name = details_soup.find("h1").text
        company_name = company_name.strip()

        all_child_divs = details_soup.findAll('div', attrs={"class": "col-md-3"})
        try:
            address = all_child_divs[1].find("p").text
        except:
            address = "Not found"
        try:
            phone_number = str(all_child_divs[2].find("p").next).replace("Telefon", "")
        except:
            phone_number = "Not Found"
        try:
            email = all_child_divs[2].find("a", attrs={"class": "mail"}).text.replace("--at--", "@")
        except:
            email = "Not Found"

        ws_email.append([company_name, email])
        ws_contact.append([company_name, address, phone_number])
        print(f"persed tata for {company_name} company ")

    offset+=100
    workbook_email.save("Company_email.xlsx")
    contact_workbook.save("Contact_details.xlsx")

workbook_email.close()
workbook_email.save("only_email.xlsx")

contact_workbook.close()
contact_workbook.save("Contact_details.xlsx")
